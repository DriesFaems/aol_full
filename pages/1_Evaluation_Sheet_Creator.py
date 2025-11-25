import io
from datetime import datetime
from typing import Dict, List

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from utils import set_branding

# Apply Branding
set_branding()

COMPETENCIES: Dict[str, str] = {
    "Ethical Decision-Making": (
        "Recognizes ethical dilemmas and applies integrity-driven reasoning "
        "in organizational and societal contexts."
    ),
    "Entrepreneurial Initiative": (
        "Identifies opportunities and mobilizes resources creatively to create economic and social value."
    ),
    "Sustainability Orientation": (
        "Integrates environmental, social, and economic considerations into decision-making and design."
    ),
    "AI-Enabled Problem Solving": (
        "Uses AI tools to analyze complex challenges and design innovative, human-supervised solutions."
    ),
    "Systems Thinking": (
        "Understands interconnections between business, society, and the environment to manage long-term consequences."
    ),
    "Democratic Leadership": (
        "Encourages inclusion, dialogue, and shared responsibility to strengthen democratic values."
    ),
    "Curiosity and Lifelong Learning": (
        "Seeks new knowledge, embraces uncertainty, and adapts to emerging technologies and trends."
    ),
    "Collaborative Intelligence": (
        "Blends human creativity, emotional intelligence, and machine intelligence to enhance performance."
    ),
    "Responsible Change Management": (
        "Leads transformation initiatives that align with ethical, social, and environmental principles."
    ),
    "Global Citizenship": (
        "Acts with cultural awareness and social responsibility to foster equitable, sustainable development."
    ),
}

RATING_OPTIONS = ["Very good", "Good", "Satisfactory", "Unsatisfactory"]
RATING_PLACEHOLDERS = {
    "Very good": "Describe what consistently outstanding performance looks like.",
    "Good": "Describe solid performance that meets expectations.",
    "Satisfactory": "Describe minimally acceptable performance.",
    "Unsatisfactory": "Describe performance that requires immediate improvement.",
}
REQUIRED_PARTICIPANT_COLUMNS = ["First name", "Last name", "Email address", "Groups"]


@st.cache_data(show_spinner=False)
def load_participants(file_buffer) -> pd.DataFrame:
    """Load the participant sheet and ensure required columns exist."""
    df = pd.read_excel(file_buffer)
    missing_columns = [
        column for column in REQUIRED_PARTICIPANT_COLUMNS if column not in df.columns
    ]
    if missing_columns:
        raise ValueError(
            f"The uploaded file is missing the following columns: {', '.join(missing_columns)}."
        )
    return df[REQUIRED_PARTICIPANT_COLUMNS].copy()


def validate_dimensions(dimensions: List[Dict]) -> List[str]:
    errors: List[str] = []
    total_weight = 0
    for index, dimension in enumerate(dimensions):
        prefix = f"Dimension {index + 1}"
        name = dimension["name"].strip()
        if not name:
            errors.append(f"{prefix}: please provide a name.")
        weight = dimension["weight"]
        if weight is None:
            errors.append(f"{prefix}: please specify a weight between 0 and 100.")
        elif weight < 0 or weight > 100:
            errors.append(f"{prefix}: weight must be between 0 and 100.")
        else:
            total_weight += weight
        for rating_label in RATING_OPTIONS:
            if not dimension["definitions"][rating_label].strip():
                errors.append(f"{prefix}: add a definition for '{rating_label}'.")
    names = [dimension["name"].strip().lower() for dimension in dimensions if dimension["name"].strip()]
    if len(names) != len(set(names)):
        errors.append("Please ensure each dimension has a unique name.")
    if not errors and total_weight != 100:
        errors.append(f"The total weight across all dimensions must equal 100. Current total: {total_weight}.")
    return errors


def build_workbook(
    participants_df: pd.DataFrame,
    assignment_weight: int,
    dimensions: List[Dict],
) -> io.BytesIO:
    roster_df = participants_df.copy()
    for dimension in dimensions:
        roster_df[f"{dimension['name']} - Score"] = ""
        roster_df[f"{dimension['name']} - Notes"] = ""
    roster_df["Final Assignment Grade"] = ""

    buffer = io.BytesIO()
    generation_time = datetime.now().strftime("%Y-%m-%d %H:%M")

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        info_df = pd.DataFrame(
            [
                {"Item": "Assignment weight (%)", "Value": assignment_weight},
                {"Item": "Number of participants", "Value": len(participants_df)},
                {"Item": "Number of dimensions", "Value": len(dimensions)},
                {"Item": "Generated on", "Value": generation_time},
            ]
        )
        info_df.to_excel(writer, sheet_name="Assignment Info", index=False)

        dimension_overview = pd.DataFrame(
            [
                {
                    "Dimension": dimension["name"],
                    "Weight (%)": dimension["weight"],
                    "Competencies covered": ", ".join(dimension["competencies"]),
                }
                for dimension in dimensions
            ]
        )
        start_row = len(info_df) + 2
        dimension_overview.to_excel(
            writer, sheet_name="Assignment Info", index=False, startrow=start_row
        )

        rubric_rows = [
            {
                "Dimension": dimension["name"],
                "Weight (%)": dimension["weight"],
                "Competencies included": "\n".join(dimension["competencies"]),
                **{rating: dimension["definitions"][rating] for rating in RATING_OPTIONS},
            }
            for dimension in dimensions
        ]
        rubric_df = pd.DataFrame(rubric_rows)
        rubric_df.to_excel(writer, sheet_name="Rubric", index=False)

        roster_df.to_excel(writer, sheet_name="Evaluation Form", index=False)

        workbook = writer.book
        evaluation_sheet = workbook["Evaluation Form"]

        # Add formulas for final assignment grade using the weighted scores
        final_grade_column_index = roster_df.columns.get_loc("Final Assignment Grade") + 1
        score_columns = [
            roster_df.columns.get_loc(f"{dimension['name']} - Score") + 1 for dimension in dimensions
        ]
        for row in range(2, len(roster_df) + 2):
            terms = []
            for dimension, column_index in zip(dimensions, score_columns):
                column_letter = get_column_letter(column_index)
                terms.append(f"({column_letter}{row}*{dimension['weight']})/100")
            formula = "+".join(terms) if terms else ""
            if formula:
                cell = evaluation_sheet.cell(row=row, column=final_grade_column_index)
                cell.value = f"={formula}"

        rubric_sheet = workbook["Rubric"]
        for column in rubric_sheet.columns:
            for cell in column:
                cell.alignment = cell.alignment.copy(wrap_text=True)

    buffer.seek(0)
    return buffer


def main() -> None:
    st.title("AoL Assignment Evaluation Builder")
    st.write(
        "Upload the course roster, define the assessment dimensions, and download a ready-to-use Excel evaluation form."
    )
    with st.expander("Competencies you can reference"):
        for name, description in COMPETENCIES.items():
            st.markdown(f"**{name}** — {description}")

    uploaded_file = st.file_uploader(
        "Upload the participant list (.xlsx)", type=["xlsx"], help="Use the provided template structure."
    )

    participants_df = None
    if uploaded_file is not None:
        try:
            participants_df = load_participants(uploaded_file)
            st.success(f"Loaded {len(participants_df)} participants.")
            st.dataframe(participants_df.head(), use_container_width=True)
        except ValueError as error:
            st.error(str(error))

    st.divider()

    with st.form("configuration_form"):
        assignment_weight = st.slider(
            "Assignment weight towards the final grade (%)", min_value=0, max_value=100, value=50, step=1
        )
        dimension_count = int(
            st.number_input(
                "How many dimensions do you want to evaluate?",
                min_value=1,
                max_value=10,
                value=2,
                step=1,
                format="%d",
            )
        )

        dimensions: List[Dict] = []
        even_split = 100 / dimension_count
        for index in range(dimension_count):
            st.subheader(f"Dimension {index + 1}")
            name = st.text_input("Name", value=f"Dimension {index + 1}", key=f"name_{index}")
            default_weight = int(round(even_split))
            weight = int(
                st.number_input(
                    "Weight (%)",
                    min_value=0,
                    max_value=100,
                    value=default_weight,
                    step=1,
                    key=f"weight_{index}",
                    help="Ensure the sum of all dimension weights equals 100%.",
                )
            )
            competencies = st.multiselect(
                "Select the competencies measured in this dimension",
                options=list(COMPETENCIES.keys()),
                key=f"competencies_{index}",
            )

            definitions: Dict[str, str] = {}
            for rating_label in RATING_OPTIONS:
                definitions[rating_label] = st.text_area(
                    f"What does '{rating_label}' mean for this dimension?",
                    key=f"definition_{index}_{rating_label}",
                    placeholder=RATING_PLACEHOLDERS[rating_label],
                )

            dimensions.append(
                {
                    "name": name,
                    "weight": weight,
                    "competencies": competencies,
                    "definitions": definitions,
                }
            )

        total_weight = sum(dimension["weight"] for dimension in dimensions)
        if total_weight == 100:
            st.caption("✅ Total dimension weight: 100%.")
        else:
            st.caption(f"⚠️ Current total dimension weight: {total_weight}% (must equal 100%).")

        submitted = st.form_submit_button("Generate evaluation workbook", use_container_width=True)

    if submitted:
        errors: List[str] = []
        if participants_df is None:
            errors.append("Please upload a participant Excel file before generating the workbook.")

        errors.extend(validate_dimensions(dimensions))

        if errors:
            for message in errors:
                st.error(message)
            return

        workbook_buffer = build_workbook(participants_df, assignment_weight, dimensions)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        file_name = f"aol_evaluation_form_{timestamp}.xlsx"
        st.success("Workbook generated successfully. Download it below.")
        st.download_button(
            "Download evaluation workbook",
            data=workbook_buffer.getvalue(),
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

if __name__ == "__main__":
    main()
