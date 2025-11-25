import io
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from utils import set_branding

# Apply Branding
set_branding()

REQUIRED_STUDENT_COLUMNS = ["First name", "Last name", "Email address", "Final Assignment Grade"]


def _read_workbook(file_bytes: bytes):
    return load_workbook(io.BytesIO(file_bytes), data_only=True)


def _parse_assignment_weight(workbook) -> float:
    sheet_name = "Assignment Info"
    if sheet_name not in workbook.sheetnames:
        raise ValueError("Missing 'Assignment Info' sheet.")
    sheet = workbook[sheet_name]
    target_label = "Assignment weight (%)"
    for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        label = row[0]
        if isinstance(label, str) and label.strip() == target_label:
            value = row[1]
            if value is None:
                break
            try:
                return float(value)
            except (TypeError, ValueError) as error:
                raise ValueError(f"Assignment weight is not numeric ({value}).") from error
    raise ValueError("Could not find 'Assignment weight (%)' entry on the Assignment Info sheet.")


def _parse_competencies(raw_value) -> List[str]:
    if raw_value is None or (isinstance(raw_value, float) and np.isnan(raw_value)):
        return []
    if isinstance(raw_value, str):
        sanitized = raw_value.replace("\n", ",")
        competencies = [part.strip() for part in sanitized.split(",") if part.strip()]
        return competencies
    if isinstance(raw_value, (list, tuple)):
        return [str(value).strip() for value in raw_value if str(value).strip()]
    return [str(raw_value).strip()]


def _parse_dimension_metadata(file_bytes: bytes) -> List[Dict]:
    try:
        rubric_df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Rubric")
    except ValueError as error:
        raise ValueError("Missing 'Rubric' sheet.") from error

    if "Dimension" not in rubric_df.columns:
        raise ValueError("Rubric sheet must contain a 'Dimension' column.")

    metadata: List[Dict] = []
    for _, row in rubric_df.iterrows():
        dimension_name = str(row["Dimension"]).strip()
        if not dimension_name or dimension_name.lower() == "dimension":
            continue
        weight_value = row.get("Weight (%)", 0)
        try:
            dimension_weight = float(weight_value)
        except (TypeError, ValueError):
            dimension_weight = 0.0
        competencies = _parse_competencies(row.get("Competencies included"))
        metadata.append(
            {
                "name": dimension_name,
                "weight": dimension_weight,
                "competencies": competencies,
            }
        )
    if not metadata:
        raise ValueError("The rubric does not contain any dimensions.")
    return metadata


def _load_evaluation_form(file_bytes: bytes) -> pd.DataFrame:
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Evaluation Form")
    except ValueError as error:
        raise ValueError("Missing 'Evaluation Form' sheet.") from error

    missing_columns = [column for column in REQUIRED_STUDENT_COLUMNS if column not in df.columns]
    if missing_columns:
        raise ValueError(f"The Evaluation Form sheet is missing required columns: {', '.join(missing_columns)}.")
    return df


def load_assignment(uploaded_file) -> Dict:
    file_bytes = uploaded_file.getvalue()
    workbook = _read_workbook(file_bytes)
    assignment_weight = _parse_assignment_weight(workbook)
    dimensions = _parse_dimension_metadata(file_bytes)
    evaluation_df = _load_evaluation_form(file_bytes)

    display_name = Path(uploaded_file.name).stem
    return {
        "source_name": uploaded_file.name,
        "display_name": display_name,
        "assignment_weight": assignment_weight,
        "dimensions": dimensions,
        "scores": evaluation_df,
    }


def merge_student_frames(left: pd.DataFrame, right: pd.DataFrame) -> pd.DataFrame:
    if left is None:
        return right
    merged = pd.merge(left, right, on="Email address", how="outer", suffixes=("", "_dup"))
    for column in ["First name", "Last name"]:
        dup_column = f"{column}_dup"
        if dup_column in merged.columns:
            merged[column] = merged[column].combine_first(merged[dup_column])
            merged = merged.drop(columns=[dup_column])
    return merged


def compute_final_grades(assignments: List[Dict]) -> pd.DataFrame:
    frames: List[pd.DataFrame] = []
    contribution_spec: List[Tuple[str, float]] = []

    for assignment in assignments:
        df = assignment["scores"].copy()
        grade_column = "Final Assignment Grade"
        assignment_label = f"{assignment['display_name']} ({assignment['assignment_weight']:.1f}%)"
        frame = df[["Email address", "First name", "Last name", grade_column]].copy()
        frame[assignment_label] = pd.to_numeric(frame[grade_column], errors="coerce")
        frame = frame.drop(columns=[grade_column])
        frames.append(frame)
        contribution_spec.append((assignment_label, assignment["assignment_weight"]))

    combined = None
    for frame in frames:
        combined = merge_student_frames(combined, frame)
    if combined is None:
        return pd.DataFrame()

    combined = combined.sort_values(by=["Last name", "First name"]).reset_index(drop=True)

    combined["Final Weighted Grade"] = 0.0
    for column_name, weight in contribution_spec:
        combined[column_name] = combined[column_name].fillna(0)
        combined["Final Weighted Grade"] += combined[column_name] * (weight / 100)

    return combined


def compute_competency_scores(assignments: List[Dict], student_names: pd.DataFrame) -> pd.DataFrame:
    competency_sums: Dict[str, Dict[str, float]] = {}
    competency_weights: Dict[str, Dict[str, float]] = {}
    used_competencies: set = set()

    for assignment in assignments:
        df = assignment["scores"]
        assignment_factor = assignment["assignment_weight"] / 100
        if assignment_factor == 0:
            continue
        for dimension in assignment["dimensions"]:
            competency_list = list(dict.fromkeys(dimension.get("competencies", [])))
            if not competency_list:
                continue
            dimension_column = f"{dimension['name']} - Score"
            if dimension_column not in df.columns:
                continue
            dimension_scores = pd.to_numeric(df[dimension_column], errors="coerce")
            dimension_factor = dimension["weight"] / 100
            weight = assignment_factor * dimension_factor
            if weight == 0:
                continue

            for competence in competency_list:
                used_competencies.add(competence)

            for email, score in zip(df["Email address"], dimension_scores):
                if not isinstance(email, str) or pd.isna(score):
                    continue
                email_key = email.strip().lower()
                student_sums = competency_sums.setdefault(email_key, {})
                student_weights = competency_weights.setdefault(email_key, {})
                for competence in competency_list:
                    student_sums[competence] = student_sums.get(competence, 0.0) + (score * weight)
                    student_weights[competence] = student_weights.get(competence, 0.0) + weight

    if not used_competencies:
        return pd.DataFrame()

    records = []
    for _, row in student_names.iterrows():
        email = row["Email address"]
        if not isinstance(email, str):
            continue
        email_key = email.strip().lower()
        record = {
            "Email address": row["Email address"],
            "First name": row["First name"],
            "Last name": row["Last name"],
        }
        student_sums = competency_sums.get(email_key, {})
        student_weights = competency_weights.get(email_key, {})
        for competence in used_competencies:
            numerator = student_sums.get(competence, 0.0)
            denominator = student_weights.get(competence, 0.0)
            record[competence] = numerator / denominator if denominator else np.nan
        records.append(record)

    competency_df = pd.DataFrame(records)
    if not competency_df.empty:
        order = ["First name", "Last name", "Email address"] + sorted(used_competencies)
        competency_df = competency_df[order]
    return competency_df


def compute_competency_coverage(assignments: List[Dict]) -> pd.DataFrame:
    coverage: Dict[str, float] = {}
    for assignment in assignments:
        assignment_weight = assignment["assignment_weight"]
        if assignment_weight == 0:
            continue
        for dimension in assignment["dimensions"]:
            dimension_weight = dimension.get("weight", 0)
            competencies = list(dict.fromkeys(dimension.get("competencies", [])))
            if dimension_weight == 0 or not competencies:
                continue
            contribution = assignment_weight * (dimension_weight / 100)
            for competence in competencies:
                coverage[competence] = coverage.get(competence, 0.0) + contribution

    if not coverage:
        return pd.DataFrame()

    coverage_rows = [
        {"Competency": competence, "Coverage (%)": contribution}
        for competence, contribution in sorted(coverage.items())
    ]
    return pd.DataFrame(coverage_rows)


def build_download(
    final_grades: pd.DataFrame,
    competency_scores: pd.DataFrame,
    competency_coverage: pd.DataFrame,
) -> io.BytesIO:
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        final_grades.to_excel(writer, sheet_name="Final Grades", index=False)
        if not competency_scores.empty:
            competency_scores.to_excel(writer, sheet_name="Competency Scores", index=False)
        if not competency_coverage.empty:
            competency_coverage.to_excel(writer, sheet_name="Competency Coverage", index=False)
    buffer.seek(0)
    return buffer


def main():
    st.title("AoL Assignment Score Aggregator")
    st.write(
        "Upload the completed evaluation workbooks for each assignment. "
        "We will combine them into a consolidated gradebook and compute weighted competency scores."
    )

    uploaded_files = st.file_uploader(
        "Upload one or more evaluation workbooks (.xlsx)", accept_multiple_files=True, type=["xlsx"]
    )

    if not uploaded_files:
        st.info("Upload at least one workbook to get started.")
        return

    assignments: List[Dict] = []
    for uploaded_file in uploaded_files:
        try:
            assignment = load_assignment(uploaded_file)
        except ValueError as error:
            st.error(f"{uploaded_file.name}: {error}")
            continue
        assignments.append(assignment)

    if not assignments:
        st.warning("No valid assignments to process.")
        return

    st.subheader("Assignments overview")
    overview_rows = [
        {
            "Assignment": assignment["display_name"],
            "Weight (%)": assignment["assignment_weight"],
            "Number of dimensions": len(assignment["dimensions"]),
        }
        for assignment in assignments
    ]
    overview_df = pd.DataFrame(overview_rows)
    st.dataframe(overview_df, use_container_width=True)

    final_grades = compute_final_grades(assignments)
    st.subheader("Final weighted grades")
    st.dataframe(final_grades, use_container_width=True)

    competency_scores = compute_competency_scores(
        assignments, final_grades[["First name", "Last name", "Email address"]]
    )
    if competency_scores.empty:
        st.warning("No competencies were mapped in the uploaded assignments, so no competency scores could be computed.")
    else:
        st.subheader("Competency-weighted scores")
        st.dataframe(competency_scores, use_container_width=True)

    competency_coverage = compute_competency_coverage(assignments)
    if competency_coverage.empty:
        st.info("No competency coverage summary is available because no competencies were referenced.")
    else:
        st.subheader("Competency coverage by weight")
        st.dataframe(competency_coverage, use_container_width=True)

    output_buffer = build_download(final_grades, competency_scores, competency_coverage)
    st.download_button(
        "Download consolidated Excel",
        data=output_buffer.getvalue(),
        file_name="aol_consolidated_scores.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
